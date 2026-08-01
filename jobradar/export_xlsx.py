"""Excel export — the application tracker she actually hands around.

Three sheets: Applied (the one that matters), Saved, and All matches. Applied is
sorted by when she marked it, so it reads as a running log of the search.
"""
from __future__ import annotations

import io
import json
from datetime import datetime, timezone

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from .db import Store

HEAD_FILL = PatternFill("solid", fgColor="F5D9C4")   # warm peach, matches the dashboard
HEAD_FONT = Font(bold=True, color="4A3428")

COLUMNS = [
    ("Applied on", 14), ("Score", 7), ("Role", 44), ("Company", 22),
    ("Location", 26), ("Source", 16), ("Posted", 12), ("Notes", 30), ("Link", 52),
]


def _reasons(row) -> str:
    try:
        return " · ".join(json.loads(row["reasons"] or "[]")[:3])
    except Exception:
        return ""


def _sheet(wb, title: str, rows, first_col_field: str):
    ws = wb.create_sheet(title)
    ws.append([c[0] for c in COLUMNS])
    for i, (name, width) in enumerate(COLUMNS, start=1):
        ws.column_dimensions[get_column_letter(i)].width = width
        cell = ws.cell(row=1, column=i)
        cell.fill = HEAD_FILL
        cell.font = HEAD_FONT
        cell.alignment = Alignment(vertical="center")
    ws.freeze_panes = "A2"

    for r in rows:
        stamp = (r[first_col_field] or "")[:10] if first_col_field in r.keys() else ""
        ws.append([
            stamp,
            round(r["score"] or 0),
            r["title"] or "",
            r["company"] or "",
            r["location"] or ("Remote" if r["remote"] else ""),
            r["source"] or "",
            (r["posted_at"] or "")[:10],
            r["notes"] or _reasons(r),
            r["url"] or "",
        ])
    # Make the link column clickable.
    for row_i in range(2, ws.max_row + 1):
        cell = ws.cell(row=row_i, column=len(COLUMNS))
        if cell.value:
            cell.hyperlink = cell.value
            cell.font = Font(color="2F6FB2", underline="single")
    return ws


def build(store: Store) -> bytes:
    wb = Workbook()
    wb.remove(wb.active)

    applied = store.conn.execute(
        "SELECT * FROM jobs WHERE status='applied' ORDER BY COALESCE(status_at,'') DESC"
    ).fetchall()
    saved = store.conn.execute(
        "SELECT * FROM jobs WHERE status='saved' ORDER BY score DESC").fetchall()
    every = store.conn.execute(
        "SELECT * FROM jobs WHERE status != 'dismissed' ORDER BY score DESC LIMIT 2000"
    ).fetchall()

    _sheet(wb, "Applied", applied, "status_at")
    _sheet(wb, "Saved", saved, "status_at")
    _sheet(wb, "All matches", every, "first_seen")

    meta = wb.create_sheet("About")
    meta["A1"] = "Job Radar export"
    meta["A1"].font = Font(bold=True, size=13)
    meta["A2"] = f"Generated {datetime.now(timezone.utc).strftime('%Y-%m-%d %H:%M UTC')}"
    meta["A3"] = f"Applied: {len(applied)}   Saved: {len(saved)}   Total matches: {len(every)}"
    meta.column_dimensions["A"].width = 60

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
